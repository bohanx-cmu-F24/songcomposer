
import json
from openpyxl import Workbook
from finetune.utils import gen_midi
import numpy as np
with open("generated_data.json", "r+") as f:
    data_in  = json.load(f)
from scipy.stats import ttest_ind, f_oneway, mannwhitneyu, kruskal
from scipy.stats import ttest_1samp, wilcoxon
import pandas as pd


joy = data_in["joy"]

class dataProcessor:
    def __init__(self,data_in):
        self.data = data_in
        self.length = {emotion: len(self.data[emotion]) for emotion in self.data}

    def computeSongLength(self):
        note_duration = {emotion: sum([sum(x["notes_duration"]) for x in self.data[emotion]]) for emotion in self.data}
        rest_duration = {emotion: sum([sum(x["rest_duration"]) for x in self.data[emotion]]) for emotion in self.data}

        # Compute "all" values
        all_note_duration = sum(note_duration.values())
        all_rest_duration = sum(rest_duration.values())

        note_duration["all"] = all_note_duration
        rest_duration["all"] = all_rest_duration

        self.note_duration = note_duration
        self.rest_duration = rest_duration

        self.note_duration_per_song = {emotion: [sum(x["notes_duration"]) for x in self.data[emotion]] for emotion in self.data}
        self.rest_duration_per_song = {emotion: [sum(x["rest_duration"]) for x in self.data[emotion]] for emotion in self.data}

        self.average_note_duration = {
            emotion: note_duration[emotion] / self.length[emotion] for emotion in self.data
        }
        self.average_note_duration["all"] = all_note_duration / sum(self.length.values())

        self.average_rest_duration = {
            emotion: rest_duration[emotion] / self.length[emotion] for emotion in self.data
        }
        self.average_rest_duration["all"] = all_rest_duration / sum(self.length.values())

        return note_duration, rest_duration

    def computePercent(self):
        self.note_percent = {
            emotion: {
                "note_percent": self.note_duration[emotion] / (
                            self.note_duration[emotion] + self.rest_duration[emotion]),
                "rest_percent": self.rest_duration[emotion] / (
                            self.note_duration[emotion] + self.rest_duration[emotion])
            }
            for emotion in self.data
        }
        self.note_percent_by_song = {
            emotion:
            [
                self.note_duration_per_song[emotion][i] / (self.note_duration_per_song[emotion][i] + self.rest_duration_per_song[emotion][i])
                for i in range(len(self.data[emotion]))
            ]
            for emotion in self.data
        }

        # Compute "all" values
        total_note_duration = self.note_duration["all"]
        total_rest_duration = self.rest_duration["all"]
        total_duration = total_note_duration + total_rest_duration

        self.note_percent["all"] = {
            "note_percent": total_note_duration / total_duration,
            "rest_percent": total_rest_duration / total_duration
        }

        return self.note_percent

    def computeNumWord(self):
        self.numWord = {
            emotion: sum([len(x["lyrics"]) for x in self.data[emotion]]) for emotion in self.data
        }
        self.numWord["all"] = sum(self.numWord.values())

        self.numWordPerSong = {
            emotion: [len(x["lyrics"]) for x in self.data[emotion]] for emotion in self.data
        }
        self.averageWord = {
            emotion: self.numWord[emotion] / self.length[emotion] for emotion in self.data
        }
        self.averageWord["all"] = self.numWord["all"] / sum(self.length.values())

        self.averageWordPerTime = {
            emotion: self.averageWord[emotion] / (
                        self.average_note_duration[emotion] + self.average_rest_duration[emotion])
            for emotion in self.data
        }

        self.averageWordPerTime["all"] = self.averageWord["all"] / (
                self.average_note_duration["all"] + self.average_rest_duration["all"]
        )

        self.averageWordPerTimeBysong = {
            emotion: [
                self.numWordPerSong[emotion][i] / (self.note_duration_per_song[emotion][i] + self.rest_duration_per_song[emotion][i])
                for i in range(len(self.data[emotion]))
            ]
            for emotion in self.data
        }

        return self.numWord, self.averageWord, self.averageWordPerTime

    def computePitches(self):
        self.averagePitch = {
            emotion: np.mean([pitch for song in self.data[emotion] for pitch in song['pitch']])
            for emotion in self.data
        }

        self.averagePitchBySong = {
            emotion: [np.mean(song['pitch']) for song in self.data[emotion]]
            for emotion in self.data
        }

        self.averagePitch["all"] = np.mean([
            pitch for emotion in self.data for song in self.data[emotion] for pitch in song['pitch']
        ])

        self.pitchRangeBySong  = {
            emotion: [max(song['pitch']) - min(song['pitch']) for song in self.data[emotion]]
            for emotion in self.data
        }
        self.averagePitchRange = {
            emotion: np.mean([max(song['pitch']) - min(song['pitch']) for song in self.data[emotion]])
            for emotion in self.data
        }
        self.averagePitchRange["all"] = np.mean([
            max(song['pitch']) - min(song['pitch']) for emotion in self.data for song in self.data[emotion]
        ])

        return self.averagePitch, self.averagePitchRange

    def exportToExcel(self, file_name: str = "output.xlsx"):
        """
        Exports all computed statistics to an Excel file.

        Args:
            file_name (str): The name of the Excel file to save the data.
        """
        wb = Workbook()

        # Add note and rest durations
        ws1 = wb.active
        ws1.title = "Durations"
        ws1.append(["Emotion", "Note Duration", "Rest Duration", "Average Note", "Average Rest"])
        for emotion in self.note_duration:
            ws1.append([
                emotion,
                self.note_duration.get(emotion, 0),
                self.rest_duration.get(emotion, 0),
                self.average_note_duration.get(emotion, 0),
                self.average_rest_duration.get(emotion, 0),
            ])

        # Add percentages
        ws2 = wb.create_sheet(title="Percentages")
        ws2.append(["Emotion", "Note Percent", "Rest Percent"])
        for emotion, values in self.note_percent.items():
            ws2.append([
                emotion,
                values["note_percent"],
                values["rest_percent"],
            ])

        # Add word statistics
        ws3 = wb.create_sheet(title="Word Counts")
        ws3.append(["Emotion", "Total Words", "Average Words", "Average words per second"])
        for emotion in self.numWord:
            ws3.append([
                emotion,
                self.numWord.get(emotion, 0),
                self.averageWord.get(emotion, 0),
                self.averageWordPerTime.get(emotion, 0)
            ])

        # Add word statistics

        ws3 = wb.create_sheet(title="Pitches")
        ws3.append(["Emotion", "Average Pitches", "Average Pitch Spread"])
        for emotion in self.averagePitch:
            ws3.append([
                emotion,
                self.averagePitch.get(emotion, 0),
                self.averagePitchRange.get(emotion, 0)
            ])

        # Save the Excel file
        wb.save(file_name)

    def test_pitches(self):
        results = {}

        all_values = [
            value for emotion in self.data for song in self.data[emotion] for value in song["pitch"]
        ]
        all_mean = np.mean(all_values)

        for emotion in self.data:
            if emotion == "all":
                continue
            stat, p_value = ttest_1samp(
                [value for song in self.data[emotion] for value in song["pitch"]],
                popmean=all_mean
            )
            results[emotion] = {
                "statistic": stat,
                "p_value": p_value,
                "significant": p_value < 0.05
            }
        return results

    def test_average_pitch_range(self):
        results = {}
        for emotion in self.data:
            if emotion == "all":
                continue
            stat, p_value = ttest_1samp(
                self.pitchRangeBySong[emotion],
                popmean=self.averagePitchRange["all"]
            )
            results[emotion] = {
                "statistic": stat,
                "p_value": p_value,
                "significant": p_value < 0.05
            }
        return results

    def test_average_word_per_sec(self):
        results = {}
        for emotion in self.data:
            if emotion == "all":
                continue
            stat, p_value = ttest_1samp(
                self.averageWordPerTimeBysong[emotion],
                popmean=self.averageWordPerTime["all"]
            )
            results[emotion] = {
                "statistic": stat,
                "p_value": p_value,
                "significant": p_value < 0.05
            }
        return results

    def test_average_note_percent(self):
        results = {}
        for emotion in self.data:
            if emotion == "all":
                continue
            stat, p_value = ttest_1samp(
                self.note_percent_by_song[emotion],
                popmean=self.note_percent["all"]["note_percent"]
            )
            results[emotion] = {
                "statistic": stat,
                "p_value": p_value,
                "significant": p_value < 0.05
            }
        return results


if __name__ == '__main__':
    processor = dataProcessor(data_in)

    # Compute statistics
    processor.computeSongLength()
    processor.computePercent()
    processor.computeNumWord()
    processor.computePitches()

    results1 = processor.test_pitches()
    print(results1)
    results2 = processor.test_average_pitch_range()
    print(results2)
    results3 = processor.test_average_word_per_sec()
    print(results3)
    results4 = processor.test_average_note_percent()
    print(results4)

    #Choose the result you care
    df = pd.DataFrame.from_dict(results4, orient='index')
    # Export the DataFrame to an Excel file
    df.to_excel('average_note_percent_t_test.xlsx')

    # Export to Excel
    # processor.exportToExcel("song_statistics.xlsx")
